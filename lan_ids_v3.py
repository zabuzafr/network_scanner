#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LAN scanner + IDS léger (sans nmap)
- Scan ARP (Scapy) + enrichissement (MAC type/vendor, reverse DNS)
- Détection d'OS best-effort via TTL (ICMP puis TCP SYN)
- Tableau console (rich si dispo, sinon ASCII)
- Exports CSV/JSON (+ XLSX si openpyxl)
- IDS passif : ARP/DNS/DHCP + scans (SYN/ICMP)
- Notifications (facultatives) : Telegram / Discord / WhatsApp Business Cloud API
- Marquage couleur persistant :
    * NEW (vert)   : IP découverte depuis < 3h
    * DOWN (rouge) : hôte connu absent du scan courant
    * SCANNER (orange) : hôte qui "frappe" récemment (scan)
    * ALERT (rouge) : hôte impliqué dans une alerte (ex: ARP spoof validé)

Utilisation rapide :
  pip install scapy manuf rich openpyxl pyyaml requests
  sudo python3 lan_ids_v3.py --config config.yaml

"""

import os, sys, csv, json, time, socket, argparse, threading
from collections import defaultdict, deque
from datetime import datetime, timedelta
from ipaddress import ip_network, ip_address

from scapy.all import (  # type: ignore
    ARP, Ether, ICMP, IP, TCP, UDP,
    DNS, DNSQR, DNSRR,
    BOOTP, DHCP,
    srp, sr1, sr, sniff, conf
)

# --- OUI/Vendor (facultatif) ---
try:
    from manuf import manuf  # type: ignore
    _manuf = manuf.MacParser()
except Exception:
    _manuf = None

# --- Console & couleurs ---
try:
    from rich.console import Console  # type: ignore
    from rich.table import Table      # type: ignore
    from rich.panel import Panel      # type: ignore
    _console = Console()
except Exception:
    _console = None

# --- XLSX (facultatif) ---
try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    _xlsx_ok = True
except Exception:
    _xlsx_ok = False

# --- YAML config (facultatif) ---
try:
    import yaml  # type: ignore
    _yaml_ok = True
except Exception:
    _yaml_ok = False

# --- Notifications externes (facultatives) ---
import requests  # type: ignore

SEV_INFO  = "INFO"
SEV_WARN  = "WARN"
SEV_ALERT = "ALERT"
_SEV_RANK = {SEV_INFO: 1, SEV_WARN: 2, SEV_ALERT: 3}

class TelegramNotifier:
    def __init__(self, token: str, chat_id: str, parse_mode: str = "Markdown"):
        self.url = f"https://api.telegram.org/bot{token}/sendMessage"
        self.chat_id = chat_id
        self.parse_mode = parse_mode
    def send(self, text: str):
        try:
            requests.post(self.url, data={
                "chat_id": self.chat_id,
                "text": text,
                "parse_mode": self.parse_mode,
                "disable_web_page_preview": True
            }, timeout=5)
        except Exception:
            pass

class DiscordNotifier:
    def __init__(self, webhook_url: str):
        self.url = webhook_url
    def send(self, text: str):
        try:
            requests.post(self.url, json={"content": text[:2000]}, timeout=5)
        except Exception:
            pass

class WhatsAppNotifier:
    """WhatsApp Business Cloud API (officiel Meta).
    - access_token: jeton d'appli (Bearer)
    - phone_number_id: ID du numéro WA Business
    - to_msisdn: numéro destinataire au format international, ex: '33612345678'
    Optionnel: template_name/template_lang pour fallback hors fenêtre de 24h.
    """
    def __init__(self, access_token: str, phone_number_id: str, to_msisdn: str,
                 api_version: str = "v20.0", template_name: str | None = None,
                 template_lang: str = "fr"):
        self.url = f"https://graph.facebook.com/{api_version}/{phone_number_id}/messages"
        self.token = access_token
        self.to = to_msisdn
        self.template_name = template_name
        self.template_lang = template_lang
    def _post(self, payload: dict):
        headers = {"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"}
        return requests.post(self.url, headers=headers, json=payload, timeout=5)
    def send_text(self, text: str):
        return self._post({
            "messaging_product": "whatsapp",
            "to": self.to,
            "type": "text",
            "text": {"body": text}
        })
    def send_template(self, title: str, detail: str):
        if not self.template_name:
            return None
        return self._post({
            "messaging_product": "whatsapp",
            "to": self.to,
            "type": "template",
            "template": {
                "name": self.template_name,
                "language": {"code": self.template_lang},
                "components": [{
                    "type": "body",
                    "parameters": [
                        {"type":"text","text": title},
                        {"type":"text","text": detail}
                    ]
                }]
            }
        })
    def send(self, text: str):
        # Essaie d'abord le texte libre
        r = None
        try:
            r = self.send_text(text)
            if r is not None and r.status_code in (200, 201):
                return
        except Exception:
            pass
        # Fallback template si défini
        if self.template_name:
            title = text.split("\n", 1)[0][:60]
            detail = text[:1000]
            try:
                self.send_template(title, detail)
            except Exception:
                pass

NOTIFIERS: list = []
MIN_NOTIFY_SEVERITY = SEV_WARN  # WARN/ALERT par défaut

def _format_alert(severity: str, event_type: str, fields: dict) -> str:
    head = f"[{severity}] {event_type}"
    if fields:
        details = "\n".join(f"- {k}={v}" for k, v in fields.items())
        return f"{head}\n{details}"
    return head

# --- Marquages persistants ---
NEW_MARK_HOURS = 3          # NEW (vert) pendant 3h
SCANNER_MARK_SEC = 3600     # SCANNER (orange) pendant 1h
ALERT_MARK_SEC   = 3600     # ALERT (rouge) pendant 1h

# ====================== Utils & Logs ======================

def ensure_files(*paths_inits: tuple[str, str | None]):
    for path, init in paths_inits:
        if not os.path.exists(path):
            with open(path, "w", encoding="utf-8") as f:
                if init:
                    f.write(init)

def rotate_jsonl(path: str, max_bytes: int = 5*1024*1024, backups: int = 5):
    if not os.path.exists(path):
        return
    if os.path.getsize(path) < max_bytes:
        return
    for i in range(backups, 0, -1):
        src = f"{path}.{i}"
        dst = f"{path}.{i+1}"
        if os.path.exists(src):
            if i == backups and os.path.exists(dst):
                os.remove(dst)
            os.replace(src, dst)
    os.replace(path, f"{path}.1")

def log_jsonl(path: str, severity: str, event_type: str, **fields):
    rotate_jsonl(path)
    record = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "severity": severity,
        "type": event_type,
        **fields
    }
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
    msg = f"{severity} [{event_type}] " + ", ".join(f"{k}={v}" for k, v in fields.items())
    if _console:
        style = "green" if severity == SEV_INFO else ("yellow" if severity == SEV_WARN else "bold red")
        _console.print(f"[{style}]{msg}[/{style}]")
    else:
        print(msg, file=sys.stderr if severity != SEV_INFO else sys.stdout)
    # Notifications externes
    if _SEV_RANK.get(severity, 0) >= _SEV_RANK.get(MIN_NOTIFY_SEVERITY, 2):
        text = _format_alert(severity, event_type, fields)
        for n in NOTIFIERS:
            try:
                n.send(text)
            except Exception:
                pass

# ====================== Helpers réseau ======================

def mac_type(mac: str) -> str:
    try:
        first_octet = int(mac.split(":")[0], 16)
    except Exception:
        return "inconnue"
    is_multicast = bool(first_octet & 0x01)
    is_local     = bool(first_octet & 0x02)
    if is_multicast:
        return "multicast"
    return "unicast, local" if is_local else "unicast, global"

def mac_vendor(mac: str) -> str | None:
    if not mac:
        return None
    if _manuf:
        try:
            return _manuf.get_manuf_long(mac)
        except Exception:
            return None
    return None

def reverse_dns(ip: str) -> str | None:
    try:
        name, _, _ = socket.gethostbyaddr(ip)
        return name
    except Exception:
        return None

# Scan ARP

def arp_scan(network_cidr, iface=None, timeout=2):
    net = ip_network(network_cidr, strict=False)
    arp = ARP(pdst=str(net))
    ether = Ether(dst="ff:ff:ff:ff:ff:ff")
    answers, _ = srp(ether/arp, timeout=timeout, retry=1, iface=iface or conf.iface, verbose=False)
    hosts = []
    for _, rcv in answers:
        hosts.append({"ip": rcv.psrc, "mac": rcv.hwsrc})
    return hosts

def arp_probe(ip: str, iface=None, timeout=1.0) -> str | None:
    ether = Ether(dst="ff:ff:ff:ff:ff:ff")
    reply, _ = srp(ether/ARP(pdst=ip), timeout=timeout, iface=iface or conf.iface, verbose=False)
    for _, rcv in reply:
        return rcv.hwsrc
    return None

# OS best-effort via TTL

def guess_os_by_ttl(ttl: int | None) -> str | None:
    if ttl is None:
        return None
    if ttl <= 70:
        return "Linux/macOS/Unix (TTL≈64)"
    if ttl <= 140:
        return "Windows (TTL≈128)"
    return "Équipement réseau/embarqué (TTL≈255)"

def os_fingerprint(ip: str, icmp_timeout=1.0, tcp_timeout=1.0, tcp_probes=(443, 80)) -> tuple[str | None, int | None, str]:
    try:
        r = sr1(IP(dst=ip)/ICMP(), timeout=icmp_timeout, verbose=False)
        if r and r.haslayer(IP):
            ttl = int(r.getlayer(IP).ttl)
            return (guess_os_by_ttl(ttl), ttl, "ICMP")
    except PermissionError:
        pass
    except Exception:
        pass
    for port in tcp_probes:
        try:
            ans, _ = sr(IP(dst=ip)/TCP(dport=port, flags="S"), timeout=tcp_timeout, verbose=False)
            for _, rcv in ans:
                if rcv.haslayer(IP):
                    ttl = int(rcv.getlayer(IP).ttl)
                    return (guess_os_by_ttl(ttl), ttl, f"TCP:{port}")
        except PermissionError:
            break
        except Exception:
            continue
    return (None, None, "none")

# ====================== IDS passif ======================

class IDS:
    """
    IDS léger :
    - ARP spoofing/flood (+ validation active ARP)
    - DNS spoofing (serveurs non autorisés + contradictions)
    - DHCP rogue (OFFER/ACK depuis serveur non autorisé)
    - SYN scan / ICMP sweep
    - Marquages temporaires warn/alert pour coloration persistante
    """
    def __init__(self, iface, alerts_path,
                 dns_allow=None, dhcp_allow=None,
                 win_scan_sec=30, syn_threshold=30, icmp_threshold=30,
                 arp_change_alert=True, arp_flood_threshold=100,
                 dns_contradiction_sec=60,
                 enable_arp_validation=True):
        self.iface = iface
        self.alerts_path = alerts_path

        self.dns_allow  = set(dns_allow or [])
        self.dhcp_allow = set(dhcp_allow or [])

        # ARP state
        self.ip2mac: dict[str,str] = {}
        self.mac2ips: defaultdict[str,set] = defaultdict(set)
        self.arp_counts: defaultdict[str,int] = defaultdict(int)
        self.arp_last_reset = time.time()
        self.arp_flood_threshold = arp_flood_threshold
        self.arp_change_alert = arp_change_alert
        self.enable_arp_validation = enable_arp_validation

        # DNS memory
        self.dns_answers: dict[str, deque] = defaultdict(deque)
        self.dns_contradiction_sec = dns_contradiction_sec

        # Scan windows
        self.win_scan_sec = win_scan_sec
        self.syn_threshold = syn_threshold
        self.icmp_threshold = icmp_threshold
        self.syn_map: dict[str, dict] = defaultdict(lambda: {"ports": set(), "ts": time.time()})
        self.icmp_map: dict[str, dict] = defaultdict(lambda: {"dests": set(), "ts": time.time()})

        # Marquage temporaire des IPs
        self._warn_ips: dict[str, float]  = {}  # ip -> expiry epoch (orange)
        self._alert_ips: dict[str, float] = {}  # ip -> expiry epoch (rouge)

        self._stop = threading.Event()
        self._thr: threading.Thread | None = None

    def start(self):
        self._thr = threading.Thread(target=self._sniff_loop, daemon=True)
        self._thr.start()

    def stop(self):
        self._stop.set()

    def _sniff_loop(self):
        bpf = "arp or icmp or tcp or (udp and (port 53 or port 67 or port 68))"
        sniff(prn=self._handle_pkt, filter=bpf, iface=self.iface or conf.iface, store=False, stop_filter=lambda p: self._stop.is_set())

    def _handle_pkt(self, p):
        try:
            if p.haslayer(ARP):
                self._handle_arp(p[ARP])
            elif p.haslayer(UDP) and p.haslayer(DNS) and p[DNS].qr == 1:
                self._handle_dns(p)
            elif p.haslayer(UDP) and (p[UDP].sport in (67,68) or p[UDP].dport in (67,68)) and p.haslayer(DHCP) and p.haslayer(BOOTP):
                self._handle_dhcp(p)
            elif p.haslayer(TCP) and p.haslayer(IP):
                self._handle_tcp(p[TCP], p[IP])
            elif p.haslayer(ICMP) and p.haslayer(IP):
                self._handle_icmp(p[ICMP], p[IP])
        except Exception:
            pass

    # --- marquage utils ---
    def _mark(self, bucket: dict[str, float], ip: str, ttl_sec: int):
        if not ip:
            return
        bucket[ip] = time.time() + ttl_sec
    def _purge(self):
        now = time.time()
        for d in (self._warn_ips, self._alert_ips):
            for ip, exp in list(d.items()):
                if exp < now:
                    d.pop(ip, None)
    def get_active_sets(self) -> tuple[set[str], set[str]]:
        self._purge()
        return set(self._warn_ips.keys()), set(self._alert_ips.keys())

    # ---------- ARP ----------
    def _handle_arp(self, arp: ARP):
        if arp.op != 2:
            return
        src_ip, src_mac = arp.psrc, arp.hwsrc
        # flood
        self.arp_counts[src_mac] += 1
        now = time.time()
        if now - self.arp_last_reset > 10:
            for mac, cnt in list(self.arp_counts.items()):
                if cnt >= self.arp_flood_threshold:
                    log_jsonl(self.alerts_path, SEV_WARN, "ARP_FLOOD", mac=mac, count=cnt, window_s=10)
                self.arp_counts[mac] = 0
            self.arp_last_reset = now
        prev = self.ip2mac.get(src_ip)
        if prev and prev.lower() != src_mac.lower() and self.arp_change_alert:
            sev = SEV_WARN
            detail = "change mapping"
            if self.enable_arp_validation:
                observed = arp_probe(src_ip, iface=self.iface, timeout=1.0)
                if observed and observed.lower() != src_mac.lower():
                    sev = SEV_ALERT
                    detail = f"validation mismatch (observed={observed}, claimed={src_mac})"
            log_jsonl(self.alerts_path, sev, "ARP_SPOOFING", ip=src_ip, prev_mac=prev, new_mac=src_mac, detail=detail)
            self._mark(self._alert_ips, src_ip, ALERT_MARK_SEC)
        self.ip2mac[src_ip] = src_mac
        self.mac2ips[src_mac].add(src_ip)
        if len(self.mac2ips[src_mac]) >= 5:
            log_jsonl(self.alerts_path, SEV_WARN, "ARP_MULTI_IP", mac=src_mac, ip_count=len(self.mac2ips[src_mac]))

    # ---------- DNS ----------
    def _handle_dns(self, p):
        ip = p[IP]
        src_dns = ip.src
        dns = p[DNS]
        qname = dns.qd.qname.decode(errors="ignore") if dns.qd and isinstance(dns.qd, DNSQR) else "<unknown>"
        answers = []
        for i in range(dns.ancount):
            rr = dns.an[i]
            if isinstance(rr, DNSRR) and rr.type in (1, 28):
                answers.append(str(rr.rdata))
        if self.dns_allow and src_dns not in self.dns_allow:
            log_jsonl(self.alerts_path, SEV_ALERT, "DNS_UNAUTH_RESP", server=src_dns, qname=qname, answers=answers)
            self._mark(self._alert_ips, src_dns, ALERT_MARK_SEC)
        now = time.time()
        dq = self.dns_answers[qname]
        dq.append((now, tuple(answers), src_dns))
        while dq and now - dq[0][0] > self.dns_contradiction_sec:
            dq.pop(0)
        uniq = {ans for _, ans, _ in dq}
        srcs = {s for _, _, s in dq}
        if len(uniq) >= 2 and len(srcs) >= 2:
            log_jsonl(self.alerts_path, SEV_WARN, "DNS_CONTRADICTION", qname=qname, sources=list(srcs))
            self._mark(self._warn_ips, src_dns, SCANNER_MARK_SEC)

    # ---------- DHCP ----------
    def _handle_dhcp(self, p):
        ip = p[IP] if p.haslayer(IP) else None
        if not ip:
            return
        server_ip = ip.src
        options = dict((k.decode() if isinstance(k, bytes) else k, v) for k, v in p[DHCP].options if isinstance(k, (str, bytes)))
        msg_type = options.get("message-type")
        if msg_type in ("offer", "ack"):
            if self.dhcp_allow and server_ip not in self.dhcp_allow:
                log_jsonl(self.alerts_path, SEV_ALERT, "DHCP_ROGUE", server=server_ip, msg_type=msg_type)
                self._mark(self._alert_ips, server_ip, ALERT_MARK_SEC)

    # ---------- TCP ----------
    def _handle_tcp(self, tcp: TCP, ip: IP):
        flags = tcp.flags
        syn = bool(flags & 0x02)
        ack = bool(flags & 0x10)
        if syn and not ack:
            rec = self.syn_map[ip.src]
            now = time.time()
            if now - rec["ts"] > self.win_scan_sec:
                rec["ports"] = set(); rec["ts"] = now
            if tcp.dport:
                rec["ports"].add(int(tcp.dport))
            if len(rec["ports"]) >= self.syn_threshold:
                log_jsonl(self.alerts_path, SEV_WARN, "TCP_SYN_SCAN", src=ip.src, ports=len(rec["ports"]), window_s=self.win_scan_sec)
                self._mark(self._warn_ips, ip.src, SCANNER_MARK_SEC)
                rec["ports"] = set(); rec["ts"] = now

    # ---------- ICMP ----------
    def _handle_icmp(self, icmp: ICMP, ip: IP):
        if icmp.type == 8:  # echo request
            rec = self.icmp_map[ip.src]
            now = time.time()
            if now - rec["ts"] > self.win_scan_sec:
                rec["dests"] = set(); rec["ts"] = now
            rec["dests"].add(ip.dst)
            if len(rec["dests"]) >= self.icmp_threshold:
                log_jsonl(self.alerts_path, SEV_WARN, "ICMP_SWEEP", src=ip.src, dests=len(rec["dests"]), window_s=self.win_scan_sec)
                self._mark(self._warn_ips, ip.src, SCANNER_MARK_SEC)
                rec["dests"] = set(); rec["ts"] = now

# ====================== Enrichissement & statut ======================

def enrich_hosts(basic_hosts, icmp_timeout, tcp_timeout, tcp_probes):
    out=[]
    for h in basic_hosts:
        ip_=h["ip"]; mac=h.get("mac")
        os_guess, ttl, ttl_src = os_fingerprint(ip_, icmp_timeout, tcp_timeout, tuple(tcp_probes))
        out.append({
            "ip": ip_,
            "mac": mac,
            "hostname": reverse_dns(ip_),
            "mac_type": mac_type(mac) if mac else None,
            "vendor": mac_vendor(mac),
            "os_guess": os_guess,
            "ttl": ttl,
            "ttl_src": ttl_src
        })
    return out

def merge_hosts(old_hosts, fresh_hosts):
    by_ip = {h["ip"]: h for h in old_hosts if "ip" in h}
    for h in fresh_hosts:
        ip_=h.get("ip")
        if not ip_:
            continue
        if ip_ in by_ip:
            for k,v in h.items():
                if k in ("ttl","ttl_src","os_guess"):
                    by_ip[ip_][k]=v
                elif v and not by_ip[ip_].get(k):
                    by_ip[ip_][k]=v
        else:
            # new host → initialiser first_seen/last_seen plus tard
            by_ip[ip_] = h
    return list(by_ip.values())

# first_seen / last_seen pour coloration NEW et suivi disponibilité

def _parse_iso(ts: str | None) -> datetime | None:
    if not ts:
        return None
    try:
        return datetime.fromisoformat(ts)
    except Exception:
        return None

def update_seen_timestamps(hosts: list[dict], reachable_ips: set[str]):
    now = datetime.now().isoformat(timespec="seconds")
    for h in hosts:
        if not h.get("first_seen"):
            h["first_seen"] = now
        if h.get("ip") in reachable_ips:
            h["last_seen"] = now

# Calcul du statut/couleur

def compute_status(ip: str, host: dict, reachable_ips: set[str],
                   warn_ips: set[str], alert_ips: set[str]) -> tuple[str, str | None]:
    # 1) ALERT
    if ip in alert_ips:
        return ("ALERT", "bold red")
    # 2) DOWN
    if ip not in reachable_ips:
        return ("DOWN", "bold red")
    # 3) SCANNER
    if ip in warn_ips:
        return ("SCANNER", "bold orange1")
    # 4) NEW < 3h
    first = _parse_iso(host.get("first_seen"))
    if first and (datetime.now() - first) < timedelta(hours=NEW_MARK_HOURS):
        return ("NEW", "bold green")
    return ("", None)

# ====================== Exports & Console ======================

def write_csv(hosts, path: str):
    fields = ["IP","Hostname","MAC","Type MAC","Vendor","OS","TTL","Src TTL","Status"]
    rows = []
    for h in sorted(hosts, key=lambda x: ip_address(x["ip"])):
        rows.append([
            h.get("ip"), h.get("hostname") or "", h.get("mac") or "",
            h.get("mac_type") or "", h.get("vendor") or "",
            h.get("os_guess") or "", h.get("ttl") if h.get("ttl") is not None else "",
            h.get("ttl_src") or "", h.get("status") or ""
        ])
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(fields); w.writerows(rows)

def write_xlsx(hosts, path: str):
    if not _xlsx_ok:
        return
    wb = Workbook(); ws = wb.active; ws.title="Scan"
    headers=["IP","Hostname","MAC","Type MAC","Vendor","OS","TTL","Src TTL","Status"]; ws.append(headers)
    for h in sorted(hosts, key=lambda x: ip_address(x["ip"])):
        ws.append([
            h.get("ip"), h.get("hostname") or "", h.get("mac") or "",
            h.get("mac_type") or "", h.get("vendor") or "",
            h.get("os_guess") or "", "" if h.get("ttl") is None else h.get("ttl"),
            h.get("ttl_src") or "", h.get("status") or ""
        ])
    for c in range(1,len(headers)+1): ws.column_dimensions[get_column_letter(c)].width=20
    wb.save(path)

def print_table(hosts_all: list[dict], reachable_ips: set[str],
                warn_ips: set[str], alert_ips: set[str], use_rich=True):
    cols = [("Status","status"),("IP","ip"),("Hostname","hostname"),("MAC","mac"),
            ("Type MAC","mac_type"),("Vendor","vendor"),("OS","os_guess"),
            ("TTL","ttl"),("Src TTL","ttl_src")]

    ordered = sorted(hosts_all, key=lambda x: ip_address(x["ip"]))
    rows = []
    for h in ordered:
        ip_ = h.get("ip")
        status, style = compute_status(ip_, h, reachable_ips, warn_ips, alert_ips)
        h["status"] = status
        row = [
            status,
            ip_ or "",
            h.get("hostname") or "",
            h.get("mac") or "",
            h.get("mac_type") or "",
            h.get("vendor") or "",
            h.get("os_guess") or "",
            "" if h.get("ttl") is None else str(h.get("ttl")),
            h.get("ttl_src") or "",
        ]
        rows.append((row, style))

    if use_rich and _console:
        table = Table(show_header=True, header_style="bold")
        for title,_ in cols:
            justify = "right" if title in ("TTL","Src TTL") else "left"
            table.add_column(title, justify=justify, no_wrap=False)
        for row, style in rows:
            table.add_row(*row, style=style)
        _console.print(table)
        return

    # Fallback ASCII
    headers = [c[0] for c in cols]
    maxw = {"Status":8,"IP":15,"Hostname":40,"MAC":17,"Type MAC":16,
            "Vendor":30,"OS":28,"TTL":5,"Src TTL":8}
    def trunc(s,w): s="" if s is None else str(s); return s if len(s)<=w else s[:w-1]+"…"
    widths=[]; tmp=[r for r,_ in rows]
    for i,hname in enumerate(headers):
        col = [trunc(r[i], maxw[hname]) for r in tmp] if tmp else []
        widths.append(min(max([len(hname)] + [len(v) for v in col]), maxw[hname]))
    def sep(): return "".join("+" + "-"*(w+2) for w in widths) + "+"
    def fmt(vals):
        out=[]
        for i,v in enumerate(vals):
            name=headers[i]; v=trunc(v,widths[i])
            out.append("| " + (v.rjust(widths[i]) if name in ("TTL","Src TTL") else v.ljust(widths[i])) + " ")
        return "".join(out)+"|"
    print(sep()); print(fmt(headers)); print(sep())
    for row,_ in rows:
        print(fmt(row))
    print(sep())

# ====================== CLI / Config / Notifiers ======================

def build_notifiers_from_args(args):
    notifs = []
    if args.tg_token and args.tg_chat:
        notifs.append(TelegramNotifier(args.tg_token, args.tg_chat))
    if args.discord_webhook:
        notifs.append(DiscordNotifier(args.discord_webhook))
    if args.wa_token and args.wa_phone_id and args.wa_to:
        notifs.append(WhatsAppNotifier(args.wa_token, args.wa_phone_id, args.wa_to,
                                       template_name=args.wa_template, template_lang=args.wa_template_lang))
    return notifs

def load_config(path: str) -> dict:
    if not path:
        return {}
    if not _yaml_ok:
        raise RuntimeError("PyYAML non installé (pip install pyyaml)")
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

def parse_args():
    p = argparse.ArgumentParser(description="Scanner LAN ARP (sans nmap) + tableau + CSV/XLSX + IDS (ARP/DNS/DHCP/Scan).")
    # arguments
    p.add_argument("--config", default=None, help="Fichier YAML de configuration.")
    p.add_argument("--network", default="192.168.1.0/24")
    p.add_argument("--iface", default=None)
    p.add_argument("--interval", type=int, default=60)
    p.add_argument("--csv", default="scan_report.csv")
    p.add_argument("--xlsx", default=None)
    p.add_argument("--json", default="scan_hosts.json")
    p.add_argument("--log", default="scan_results.log")
    p.add_argument("--alerts", default="alerts.jsonl")
    p.add_argument("--icmp-timeout", type=float, default=1.0)
    p.add_argument("--tcp-timeout", type=float, default=1.0)
    p.add_argument("--tcp-probes", nargs="+", type=int, default=[443, 80])
    p.add_argument("--no-clear", action="store_true")
    p.add_argument("--no-rich", action="store_true")
    p.add_argument("--once", action="store_true")
    # IDS
    p.add_argument("--no-ids", action="store_true")
    p.add_argument("--dns-servers", nargs="*", default=[])
    p.add_argument("--dhcp-servers", nargs="*", default=[])
    p.add_argument("--ids-window", type=int, default=30)
    p.add_argument("--syn-threshold", type=int, default=30)
    p.add_argument("--icmp-threshold", type=int, default=30)
    p.add_argument("--arp-flood-threshold", type=int, default=100)
    p.add_argument("--dns-contradiction-sec", type=int, default=60)
    p.add_argument("--no-arp-change-alert", action="store_true")
    p.add_argument("--no-arp-validation", action="store_true")
    # Notifications externes
    p.add_argument("--notify-min-level", default="WARN", choices=["INFO","WARN","ALERT"],
                   help="Seuil minimal pour notifier (par défaut WARN).")
    # Telegram
    p.add_argument("--tg-token", default=None)
    p.add_argument("--tg-chat",  default=None)
    # Discord
    p.add_argument("--discord-webhook", default=None)
    # WhatsApp Business Cloud
    p.add_argument("--wa-token",    default=None)
    p.add_argument("--wa-phone-id", default=None)
    p.add_argument("--wa-to",       default=None)
    p.add_argument("--wa-template", default=None, help="Nom de template WA pour fallback hors 24h (optionnel).")
    p.add_argument("--wa-template-lang", default="fr")

    # 1) parse args utilisateur
    args = p.parse_args()
    # 2) valeurs par défaut réelles
    defaults = vars(p.parse_args([]))
    setattr(args, "_defaults", defaults)
    return args

def apply_config(args, cfg: dict):
    """Applique le YAML sans écraser des options passées en CLI.
       Règle : on remplace uniquement si la valeur courante == valeur par défaut.
    """
    defaults = getattr(args, "_defaults", {})
    for k in ["network","iface","interval","csv","xlsx","json","log","alerts",
              "icmp_timeout","tcp_timeout","tcp_probes"]:
        if k in cfg and getattr(args, k, None) == defaults.get(k):
            setattr(args, k, cfg[k])
    ids = cfg.get("ids", {})
    if ids:
        if "dns_servers" in ids and args.dns_servers == defaults.get("dns_servers", []):
            args.dns_servers = ids["dns_servers"]
        if "dhcp_servers" in ids and args.dhcp_servers == defaults.get("dhcp_servers", []):
            args.dhcp_servers = ids["dhcp_servers"]
        if "window" in ids and args.ids_window == defaults.get("ids_window", 30):
            args.ids_window = ids["window"]
        if "syn_threshold" in ids and args.syn_threshold == defaults.get("syn_threshold", 30):
            args.syn_threshold = ids["syn_threshold"]
        if "icmp_threshold" in ids and args.icmp_threshold == defaults.get("icmp_threshold", 30):
            args.icmp_threshold = ids["icmp_threshold"]
        if "arp_flood_threshold" in ids and args.arp_flood_threshold == defaults.get("arp_flood_threshold", 100):
            args.arp_flood_threshold = ids["arp_flood_threshold"]
        if "dns_contradiction_sec" in ids and args.dns_contradiction_sec == defaults.get("dns_contradiction_sec", 60):
            args.dns_contradiction_sec = ids["dns_contradiction_sec"]
        if "arp_change_alert" in ids and args.no_arp_change_alert == defaults.get("no_arp_change_alert", False):
            args.no_arp_change_alert = not bool(ids["arp_change_alert"])
        if "arp_validation" in ids and args.no_arp_validation == defaults.get("no_arp_validation", False):
            args.no_arp_validation = not bool(ids["arp_validation"])
    notify = cfg.get("notify", {})
    if notify:
        if "min_level" in notify and args.notify_min_level == defaults.get("notify_min_level", "WARN"):
            args.notify_min_level = notify["min_level"]
        tg = notify.get("telegram", {})
        if tg and not (args.tg_token or args.tg_chat):
            args.tg_token = tg.get("token"); args.tg_chat = tg.get("chat")
        dc = notify.get("discord", {})
        if dc and not args.discord_webhook:
            args.discord_webhook = dc.get("webhook")
        wa = notify.get("whatsapp", {})
        if wa and not (args.wa_token or args.wa_phone_id or args.wa_to):
            args.wa_token = wa.get("access_token")
            args.wa_phone_id = wa.get("phone_number_id")
            args.wa_to = wa.get("to")
            args.wa_template = wa.get("template")
            args.wa_template_lang = wa.get("template_lang", "fr")

# ====================== Exécution ======================

def run_once(args, ids):
    ensure_files((args.log, "# Nouvelles IP détectées\n"), (args.json, None), (args.alerts, None))
    known = []
    try:
        with open(args.json, "r", encoding="utf-8") as f:
            known = json.load(f) or []
    except Exception:
        known = []

    basic = arp_scan(args.network, iface=args.iface, timeout=2)
    enriched = enrich_hosts(basic, args.icmp_timeout, args.tcp_timeout, args.tcp_probes)

    now_ips = {h["ip"] for h in enriched}
    old_ips = {h["ip"] for h in known if "ip" in h}
    new_ips = now_ips - old_ips
    if new_ips:
        msg = "Nouvelles IP: " + ", ".join(sorted(new_ips, key=lambda x: ip_address(x)))
        if _console: _console.print(f"[bold green]{msg}[/bold green]")
        else: print(msg)
        with open(args.log, "a", encoding="utf-8") as f:
            for ip_ in sorted(new_ips, key=lambda x: ip_address(x)):
                f.write(f"{datetime.now().isoformat(timespec='seconds')} - NEW IP: {ip_}\n")

    merged = merge_hosts(known, enriched)

    # Met à jour first_seen / last_seen
    reachable_ips = {h["ip"] for h in enriched}
    update_seen_timestamps(merged, reachable_ips)

    # Sauvegardes / exports (status sera ajouté au moment de l'affichage)
    with open(args.json, "w", encoding="utf-8") as f:
        json.dump(sorted(merged, key=lambda h: ip_address(h["ip"])), f, ensure_ascii=False, indent=2)
    write_csv(merged, args.csv)
    if args.xlsx:
        write_xlsx(merged, args.xlsx)

    # Récupère les IP marquées warn/alert par l’IDS
    warn_ips, alert_ips = (set(), set())
    if ids:
        warn_ips, alert_ips = ids.get_active_sets()

    # Affichage
    if not args.no_clear:
        os.system("cls" if os.name == "nt" else "clear")
    header = f"=== État du réseau au {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ==="
    if _console:
        _console.print(Panel(header))
    else:
        print(header)

    # Affiche tous les hôtes connus (y compris DOWN)
    print_table(merged, reachable_ips, warn_ips, alert_ips, use_rich=(not args.no_rich))

    footer = f"(Rapports → CSV: {args.csv}" + (f", XLSX: {args.xlsx}" if args.xlsx else "") + f", JSON: {args.json}, Alerts JSONL: {args.alerts})\n"
    if _console:
        _console.print(footer)
    else:
        print(footer)

    # Synchronise mapping ARP côté IDS
    if ids:
        for h in enriched:
            if h.get("ip") and h.get("mac"):
                prev = ids.ip2mac.get(h["ip"])
                if prev and prev.lower()!=h["mac"].lower() and not args.no_arp_change_alert:
                    if not args.no_arp_validation:
                        observed = arp_probe(h["ip"], iface=args.iface, timeout=1.0)
                        if observed and observed.lower()!=h["mac"].lower():
                            log_jsonl(args.alerts, SEV_ALERT, "ARP_SPOOFING", ip=h["ip"], prev_mac=prev, new_mac=h["mac"], detail=f"validation mismatch (observed={observed})")
                            ids._mark(ids._alert_ips, h["ip"], ALERT_MARK_SEC)
                ids.ip2mac[h["ip"]] = h["mac"]
                ids.mac2ips[h["mac"]].add(h["ip"])


def run_loop(args):
    ids = None
    if not args.no_ids:
        ids = IDS(
            iface=args.iface,
            alerts_path=args.alerts,
            dns_allow=args.dns_servers,
            dhcp_allow=args.dhcp_servers,
            win_scan_sec=args.ids_window,
            syn_threshold=args.syn_threshold,
            icmp_threshold=args.icmp_threshold,
            arp_change_alert=not args.no_arp_change_alert,
            arp_flood_threshold=args.arp_flood_threshold,
            dns_contradiction_sec=args.dns_contradiction_sec,
            enable_arp_validation=not args.no_arp_validation
        )
        ids.start()
    try:
        while True:
            run_once(args, ids)
            time.sleep(args.interval)
    finally:
        if ids: ids.stop()


def main():
    try:
        args = parse_args()
        if args.config:
            cfg = load_config(args.config)
            apply_config(args, cfg)
        # Notifiers
        global NOTIFIERS, MIN_NOTIFY_SEVERITY
        NOTIFIERS = build_notifiers_from_args(args)
        MIN_NOTIFY_SEVERITY = args.notify_min_level
        # Run
        if args.once:
            run_once(args, None if args.no_ids else IDS(
                iface=args.iface, alerts_path=args.alerts,
                dns_allow=args.dns_servers, dhcp_allow=args.dhcp_servers,
                win_scan_sec=args.ids_window, syn_threshold=args.syn_threshold, icmp_threshold=args.icmp_threshold,
                arp_change_alert=not args.no_arp_change_alert, arp_flood_threshold=args.arp_flood_threshold,
                dns_contradiction_sec=args.dns_contradiction_sec, enable_arp_validation=not args.no_arp_validation
            ))
        else:
            run_loop(args)
    except KeyboardInterrupt:
        print("\nInterruption utilisateur, arrêt propre."); sys.exit(0)
    except PermissionError:
        print("Permission refusée : exécute avec des privilèges (sudo/cap_net_raw).", file=sys.stderr); sys.exit(1)
    except Exception as e:
        print(f"Erreur : {e}", file=sys.stderr); sys.exit(1)

if __name__ == "__main__":
    main()
