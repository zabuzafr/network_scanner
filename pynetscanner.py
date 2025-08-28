#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import json
import time
import sys
import csv
import socket
import argparse
from datetime import datetime
from ipaddress import ip_network, ip_address

from scapy.all import ARP, Ether, ICMP, IP, TCP, srp, sr1, sr, conf  # type: ignore

# --- OUI / Vendor (facultatif via 'manuf') ---
try:
    from manuf import manuf  # type: ignore
    _manuf = manuf.MacParser()
except Exception:
    _manuf = None

# --- Console table (optionnel via 'rich') ---
try:
    from rich.console import Console  # type: ignore
    from rich.table import Table      # type: ignore
    _rich_console = Console()
except Exception:
    _rich_console = None

# --- XLSX (optionnel via openpyxl) ---
try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    _xlsx_ok = True
except Exception:
    _xlsx_ok = False

# ====================== Utilitaires ======================

def ensure_logs(log_txt_path: str, json_path: str):
    if not os.path.exists(log_txt_path):
        with open(log_txt_path, "w", encoding="utf-8") as f:
            f.write("# Journal des nouvelles IP détectées\n")
    if not os.path.exists(json_path):
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump([], f)

def load_known_hosts(json_path: str):
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        return []

def save_known_hosts(hosts, json_path: str):
    hosts_sorted = sorted(hosts, key=lambda h: ip_address(h["ip"]))
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(hosts_sorted, f, ensure_ascii=False, indent=2)

def append_log_new_ips(new_ips, log_txt_path: str):
    ts = datetime.now().isoformat(timespec="seconds")
    with open(log_txt_path, "a", encoding="utf-8") as f:
        for ip_ in sorted(new_ips, key=lambda x: ip_address(x)):
            f.write(f"{ts} - NEW IP: {ip_}\n")

def merge_hosts(old_hosts, fresh_hosts):
    by_ip = {h["ip"]: h for h in old_hosts if "ip" in h}
    for h in fresh_hosts:
        ip_ = h.get("ip")
        if not ip_:
            continue
        if ip_ in by_ip:
            # remplace TTL/OS (instantané), complète champs manquants
            for k, v in h.items():
                if k in ("ttl", "ttl_src", "os_guess"):
                    by_ip[ip_][k] = v
                elif v and not by_ip[ip_].get(k):
                    by_ip[ip_][k] = v
        else:
            by_ip[ip_] = h
    return list(by_ip.values())

def mac_type(mac: str) -> str:
    """Type d’adresse MAC: multicast vs unicast, global vs local."""
    try:
        first_octet = int(mac.split(":")[0], 16)
    except Exception:
        return "inconnue"
    is_multicast = bool(first_octet & 0x01)
    is_local     = bool(first_octet & 0x02)
    if is_multicast:
        return "multicast"
    return "unicast, local" if is_local else "unicast, global"

def mac_vendor(mac: str) -> str | None:
    if not mac:
        return None
    if _manuf:
        try:
            v = _manuf.get_manuf_long(mac)
            return v
        except Exception:
            return None
    return None

def reverse_dns(ip: str) -> str | None:
    try:
        name, _, _ = socket.gethostbyaddr(ip)
        return name
    except Exception:
        return None

# ====================== Scan & OS Guess ======================

def arp_scan(network_cidr, iface=None, timeout=2):
    """ARP who-has (broadcast) → liste dicts {ip, mac}"""
    net = ip_network(network_cidr, strict=False)
    arp = ARP(pdst=str(net))
    ether = Ether(dst="ff:ff:ff:ff:ff:ff")
    answers, _ = srp(
        ether/arp,
        timeout=timeout,
        retry=1,
        iface=iface or conf.iface,
        verbose=False
    )
    hosts = []
    for _, rcv in answers:
        hosts.append({"ip": rcv.psrc, "mac": rcv.hwsrc})
    return hosts

def guess_os_by_ttl(ttl: int | None) -> str | None:
    """Heuristique simple OS ← TTL observé."""
    if ttl is None:
        return None
    if ttl <= 70:
        return "Linux/macOS/Unix (TTL≈64)"
    if ttl <= 140:
        return "Windows (TTL≈128)"
    return "Équipement réseau/embarqué (TTL≈255)"

def os_fingerprint(ip: str, icmp_timeout=1.0, tcp_timeout=1.0, tcp_probes=(443, 80)) -> tuple[str | None, int | None, str]:
    """
    Best-effort:
      1) ICMP Echo → TTL
      2) TCP SYN sur ports courants → TTL
    Retourne (os_guess, ttl, method)
    """
    # 1) ICMP
    try:
        r = sr1(IP(dst=ip)/ICMP(), timeout=icmp_timeout, verbose=False)
        if r and r.haslayer(IP):
            ttl = int(r.getlayer(IP).ttl)
            return (guess_os_by_ttl(ttl), ttl, "ICMP")
    except PermissionError:
        pass
    except Exception:
        pass

    # 2) TCP
    for port in tcp_probes:
        try:
            ans, _ = sr(IP(dst=ip)/TCP(dport=port, flags="S"), timeout=tcp_timeout, verbose=False)
            for _, rcv in ans:
                if rcv.haslayer(IP):
                    ttl = int(rcv.getlayer(IP).ttl)
                    return (guess_os_by_ttl(ttl), ttl, f"TCP:{port}")
        except PermissionError:
            break
        except Exception:
            continue

    return (None, None, "none")

# ====================== Exports & Console ======================

def write_csv_report(hosts, csv_path: str):
    fields = ["IP", "Hostname", "MAC", "Type MAC", "Vendor", "OS", "TTL", "Source TTL"]
    hosts_sorted = sorted(hosts, key=lambda h: ip_address(h["ip"]))
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(fields)
        for h in hosts_sorted:
            w.writerow([
                h.get("ip"),
                h.get("hostname") or "",
                h.get("mac") or "",
                h.get("mac_type") or "",
                h.get("vendor") or "",
                h.get("os_guess") or "",
                h.get("ttl") if h.get("ttl") is not None else "",
                h.get("ttl_src") or "",
            ])

def write_xlsx_report(hosts, xlsx_path: str):
    if not _xlsx_ok:
        print("openpyxl non installé : XLSX non généré.", file=sys.stderr)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan"
    headers = ["IP", "Hostname", "MAC", "Type MAC", "Vendor", "OS", "TTL", "Source TTL"]
    ws.append(headers)
    for h in sorted(hosts, key=lambda x: ip_address(x["ip"])):
        ws.append([
            h.get("ip"),
            h.get("hostname") or "",
            h.get("mac") or "",
            h.get("mac_type") or "",
            h.get("vendor") or "",
            h.get("os_guess") or "",
            "" if h.get("ttl") is None else h.get("ttl"),
            h.get("ttl_src") or "",
        ])
    # ajustement simple des largeurs
    for col_idx in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    wb.save(xlsx_path)

def print_console_table(hosts, use_rich: bool = True):
    cols = [
        ("IP", "ip"),
        ("Hostname", "hostname"),
        ("MAC", "mac"),
        ("Type MAC", "mac_type"),
        ("Vendor", "vendor"),
        ("OS", "os_guess"),
        ("TTL", "ttl"),
        ("Src TTL", "ttl_src"),
    ]
    rows = []
    for h in sorted(hosts, key=lambda x: ip_address(x["ip"])):
        rows.append([
            h.get("ip") or "",
            h.get("hostname") or "",
            h.get("mac") or "",
            h.get("mac_type") or "",
            h.get("vendor") or "",
            h.get("os_guess") or "",
            "" if h.get("ttl") is None else str(h.get("ttl")),
            h.get("ttl_src") or "",
        ])

    if use_rich and _rich_console:
        table = Table(show_header=True, header_style="bold")
        for title, _ in cols:
            justify = "right" if title in ("TTL", "Src TTL") else "left"
            table.add_column(title, justify=justify, no_wrap=False)
        for r in rows:
            table.add_row(*r)
        _rich_console.print(table)
        return

    # Fallback ASCII
    maxw = {
        "IP": 15, "Hostname": 40, "MAC": 17, "Type MAC": 16,
        "Vendor": 30, "OS": 28, "TTL": 5, "Src TTL": 8
    }
    def trunc(s, w):
        s = "" if s is None else str(s)
        return s if len(s) <= w else s[:w-1] + "…"

    headers = [c[0] for c in cols]
    widths = []
    for hname in headers:
        idx = headers.index(hname)
        col_vals = [trunc(r[idx], maxw[hname]) for r in rows] if rows else []
        w = max([len(hname)] + [len(v) for v in col_vals])
        widths.append(min(w, maxw[hname]))

    def sep():
        parts = ["+" + "-"*(w+2) for w in widths]
        return "".join(parts) + "+"

    def fmt_row(vals):
        outs = []
        for i, v in enumerate(vals):
            hname = headers[i]
            v = trunc(v, widths[i])
            if hname in ("TTL", "Src TTL"):
                outs.append("| " + v.rjust(widths[i]) + " ")
            else:
                outs.append("| " + v.ljust(widths[i]) + " ")
        return "".join(outs) + "|"

    print(sep())
    print(fmt_row(headers))
    print(sep())
    for r in rows:
        print(fmt_row(r))
    print(sep())

# ====================== Pipeline ======================

def build_host_records(basic_hosts: list[dict], icmp_timeout, tcp_timeout, tcp_probes) -> list[dict]:
    """Complète les enregistrements avec hostname, mac_type, vendor, os_guess"""
    enriched = []
    for h in basic_hosts:
        ip_ = h["ip"]
        mac = h.get("mac")
        hostname = reverse_dns(ip_)
        mtype = mac_type(mac) if mac else None
        vendor = mac_vendor(mac)
        os_guess, ttl, ttl_src = os_fingerprint(ip_, icmp_timeout=icmp_timeout, tcp_timeout=tcp_timeout, tcp_probes=tuple(tcp_probes))
        enriched.append({
            "ip": ip_,
            "mac": mac,
            "hostname": hostname,
            "mac_type": mtype,
            "vendor": vendor,
            "os_guess": os_guess,
            "ttl": ttl,
            "ttl_src": ttl_src
        })
    return enriched

def run_once(args):
    ensure_logs(args.log, args.json)
    known_hosts = load_known_hosts(args.json)

    basic = arp_scan(args.network, iface=args.iface, timeout=2)
    enriched = build_host_records(basic, args.icmp_timeout, args.tcp_timeout, args.tcp_probes)

    current_ips = {h["ip"] for h in enriched}
    known_ips = {h["ip"] for h in known_hosts}
    new_ips = current_ips - known_ips
    if new_ips:
        print(f"Détection de {len(new_ips)} nouvelles IP :")
        for ip_ in sorted(new_ips, key=lambda x: ip_address(x)):
            print(f"- {ip_}")
        append_log_new_ips(new_ips, args.log)

    known_hosts = merge_hosts(known_hosts, enriched)
    save_known_hosts(known_hosts, args.json)
    write_csv_report(known_hosts, args.csv)
    if args.xlsx:
        write_xlsx_report(known_hosts, args.xlsx)

    if not args.no_clear:
        os.system("cls" if os.name == "nt" else "clear")
    print(f"=== État du réseau au {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
    print_console_table(enriched, use_rich=(not args.no_rich))
    print(f"(Rapports mis à jour → CSV: {args.csv}" + (f", XLSX: {args.xlsx}" if args.xlsx else "") + f", JSON: {args.json})\n")

def run_loop(args):
    while True:
        run_once(args)
        time.sleep(args.interval)

def parse_args():
    p = argparse.ArgumentParser(description="Scanner LAN ARP (sans nmap) avec affichage console, CSV/XLSX, et détection d’OS (TTL).")
    p.add_argument("--network", default="192.168.0.0/24", help="Plage réseau CIDR (ex: 192.168.1.0/24)")
    p.add_argument("--iface", default=None, help="Interface réseau (ex: eth0, en0, Wi-Fi). Défaut: interface Scapy.")
    p.add_argument("--interval", type=int, default=60, help="Intervalle (secondes) entre scans en mode boucle.")
    p.add_argument("--csv", default="scan_report.csv", help="Fichier de sortie CSV.")
    p.add_argument("--xlsx", default=None, help="Fichier de sortie XLSX (nécessite openpyxl).")
    p.add_argument("--json", default="scan_hosts.json", help="Fichier d’état JSON.")
    p.add_argument("--log", default="scan_results.log", help="Fichier journal des nouvelles IP.")
    p.add_argument("--icmp-timeout", type=float, default=1.0, help="Timeout ICMP (s).")
    p.add_argument("--tcp-timeout", type=float, default=1.0, help="Timeout TCP (s).")
    p.add_argument("--tcp-probes", nargs="+", type=int, default=[443, 80], help="Ports TCP à sonder pour TTL si ICMP bloqué.")
    p.add_argument("--no-clear", action="store_true", help="Ne pas effacer l’écran entre itérations.")
    p.add_argument("--no-rich", action="store_true", help="Désactiver l’affichage 'rich' (forcer ASCII).")
    p.add_argument("--once", action="store_true", help="Effectuer un seul scan puis quitter.")
    return p.parse_args()

def main():
    try:
        args = parse_args()
        if args.once:
            run_once(args)
        else:
            run_loop(args)
    except KeyboardInterrupt:
        print("\nInterruption utilisateur, arrêt propre.")
        sys.exit(0)
    except PermissionError:
        print("Permission refusée : exécute le script avec des privilèges élevés (sudo/administrateur).", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Erreur : {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
